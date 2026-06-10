import json
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi.responses import StreamingResponse
import logging
import os
import time
from email.header import decode_header
from datetime import datetime, timezone
from typing import Any, Dict, Optional

import requests
import uvicorn
from dotenv import load_dotenv
import re
from fastapi import FastAPI, File, Form, Header, Query, Request, UploadFile
from fastapi.responses import JSONResponse, Response
from requests import Response as RequestsResponse
from starlette.exceptions import HTTPException as StarletteHTTPException


load_dotenv()
APP_VERSION = "2026-06-05.2"


class Config:
    DOCINT_ENDPOINT = os.getenv("DOCINT_ENDPOINT", "").rstrip("/")
    DOCINT_KEY = os.getenv("DOCINT_KEY", "")
    DOCINT_API_VERSION = os.getenv("DOCINT_API_VERSION", "2024-11-30")
    DOCINT_MODEL_ID = os.getenv("DOCINT_MODEL_ID", "prebuilt-read")

    AOAI_ENDPOINT = os.getenv("AOAI_ENDPOINT", "").rstrip("/")
    AOAI_KEY = os.getenv("AOAI_KEY", "")
    AOAI_DEPLOYMENT = os.getenv("AOAI_DEPLOYMENT", "")
    AOAI_API_VERSION = os.getenv("AOAI_API_VERSION", "2024-02-01")

    HTTP_CONNECT_TIMEOUT_SEC = float(os.getenv("HTTP_CONNECT_TIMEOUT_SEC", "10"))
    HTTP_READ_TIMEOUT_SEC = float(os.getenv("HTTP_READ_TIMEOUT_SEC", "60"))
    DOCINT_POLL_MAX_RETRIES = int(os.getenv("DOCINT_POLL_MAX_RETRIES", "30"))
    DOCINT_POLL_INTERVAL_SEC = float(os.getenv("DOCINT_POLL_INTERVAL_SEC", "2"))
    DOCINT_POLL_BACKOFF_MULTIPLIER = float(os.getenv("DOCINT_POLL_BACKOFF_MULTIPLIER", "1.2"))
    DOCINT_POLL_MAX_INTERVAL_SEC = float(os.getenv("DOCINT_POLL_MAX_INTERVAL_SEC", "10"))


class APIError(Exception):
    def __init__(self, message: str, status_code: int = 500, details: Optional[Dict[str, Any]] = None) -> None:
        super().__init__(message)
        self.message = message
        self.status_code = status_code
        self.details = details or {}


setup_logging_done = False


def setup_logging() -> None:
    global setup_logging_done
    if setup_logging_done:
        return

    log_level = os.getenv("LOG_LEVEL", "INFO").upper()
    logging.basicConfig(
        level=log_level,
        format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    )
    setup_logging_done = True


app = FastAPI(title="OCR + AI Extraction API", version=APP_VERSION)
setup_logging()


@app.exception_handler(APIError)
async def api_error_handler(_: Request, err: APIError) -> JSONResponse:
    payload: Dict[str, Any] = {"error": err.message}
    if err.details:
        payload["details"] = err.details
    logging.warning("API error: %s", payload)
    return JSONResponse(status_code=err.status_code, content=payload)


@app.exception_handler(requests.Timeout)
async def timeout_handler(_: Request, err: requests.Timeout) -> JSONResponse:
    logging.exception("Timeout while calling external service")
    return JSONResponse(
        status_code=504,
        content={"error": "External service timeout.", "details": {"reason": str(err)}},
    )


@app.exception_handler(requests.RequestException)
async def request_exception_handler(_: Request, err: requests.RequestException) -> JSONResponse:
    logging.exception("Request exception while calling external service")
    return JSONResponse(
        status_code=502,
        content={"error": "External service request failed.", "details": {"reason": str(err)}},
    )


@app.exception_handler(StarletteHTTPException)
async def starlette_http_exception_handler(_: Request, err: StarletteHTTPException) -> JSONResponse:
    return JSONResponse(
        status_code=err.status_code,
        content={"error": "HTTP error", "details": {"reason": str(err.detail)}},
    )


@app.exception_handler(Exception)
async def unexpected_error_handler(_: Request, err: Exception) -> JSONResponse:
    logging.exception("Unexpected server error")
    return JSONResponse(
        status_code=500,
        content={"error": "Internal server error.", "details": {"reason": str(err)}},
    )


@app.get("/health")
async def health() -> Dict[str, str]:
    return {
        "status": "ok",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "version": APP_VERSION,
    }


def _decode_header_value(raw: Optional[str]) -> str:
    """Re-decode header value that may have been mis-decoded as latin-1 but is actually UTF-8."""
    if not raw:
        return ""
    try:
        return raw.encode("latin-1").decode("utf-8")
    except (UnicodeDecodeError, UnicodeEncodeError):
        return raw


@app.post("/process")
async def process_document(
    request: Request,
    file: UploadFile = File(...),
    prompt_file: Optional[UploadFile] = File(default=None),
    format: Optional[str] = Query(default=None),
    prompt: Optional[str] = Query(default=None),
    filename: Optional[str] = Query(default=None),
) -> Response:
    validate_required_env()

    if not file or not file.filename:
        raise APIError("No file provided.", 400)

    upload_filename = os.path.basename(file.filename)
    content_type = file.content_type or "application/octet-stream"

    if not is_allowed_file(upload_filename, content_type):
        raise APIError("Unsupported file type. Allowed: PDF or image formats.", 400)

    file_bytes = await file.read()
    if not file_bytes:
        raise APIError("Uploaded file is empty.", 400)

    # Read prompt from uploaded .txt file if provided
    txt_prompt = ""
    if prompt_file and prompt_file.filename:
        txt_ext = os.path.splitext(prompt_file.filename.lower())[1]
        if txt_ext != ".txt":
            raise APIError("prompt_file must be a .txt file.", 400)
        txt_bytes = await prompt_file.read()
        txt_prompt = txt_bytes.decode("utf-8", errors="replace").strip()

    # Read headers for format only (prompt via file or query param)
    x_format = _decode_header_value(request.headers.get("x-format"))

    # Determine output format: explicit param > filename extension > header > default
    output_filename = filename
    ext_format: Optional[str] = None
    if output_filename:
        ext = os.path.splitext(output_filename.lower())[1].lstrip(".")
        if ext in ("json", "csv", "xlsx", "txt"):
            ext_format = ext

    output_format = (format or ext_format or x_format or "xlsx").lower()

    # Priority: txt file > query param prompt
    user_prompt = (txt_prompt or prompt or "").strip()

    # Build the download filename
    if not output_filename:
        output_filename = f"result.{output_format}"

    logging.info(
        "Processing request for file=%s, content_type=%s, format=%s, output_filename=%s, prompt_source=%s",
        upload_filename, content_type, output_format, output_filename,
        "txt_file" if txt_prompt else "query_param" if prompt else "none"
    )
    logging.info("[PROMPT] user_prompt: %s", {user_prompt})

    ocr_text = run_document_intelligence_ocr(
        file_bytes=file_bytes,
        content_type=content_type
    )

    ai_result = run_aoai_extraction(ocr_text=ocr_text, prompt=user_prompt)
    decoded_filename = decode_mime_filename(file.filename)
    ai_result["ファイル参考"] = decoded_filename
    

    logging.info("ocr_text: %s", ocr_text)
    logging.info("ai_result: %s", ai_result)

    if output_format == "json":
        return Response(
            content=json.dumps(ai_result, ensure_ascii=False, indent=2),
            media_type="application/json; charset=utf-8",
        )

    elif output_format == "txt":
        lines = [f"{k}: {v}" for k, v in ai_result.items()]
        txt_content = "\n".join(lines)
        return Response(
            content=txt_content,
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={output_filename}"},
        )

    elif output_format == "csv":
        data = flatten_json(ai_result)
        csv_text = dict_to_csv_vertical(data)

        # ✅ Add UTF-8 BOM for Excel
        csv_text = "\ufeff" + csv_text

        return Response(
            content=csv_text,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={output_filename}"},
        )


    else:
        xlsx_bytes = build_xlsx_from_ai_result(ai_result)
        return StreamingResponse(
            io.BytesIO(xlsx_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={output_filename}"},
        )


def validate_required_env() -> None:
    required = {
        "DOCINT_ENDPOINT": Config.DOCINT_ENDPOINT,
        "DOCINT_KEY": Config.DOCINT_KEY,
        "AOAI_ENDPOINT": Config.AOAI_ENDPOINT,
        "AOAI_KEY": Config.AOAI_KEY,
        "AOAI_DEPLOYMENT": Config.AOAI_DEPLOYMENT,
    }
    missing = [k for k, v in required.items() if not v]
    if missing:
        raise APIError("Missing required environment variables.", 500, {"missing": missing})

def decode_mime_filename(filename: str) -> str:
    try:
        decoded_parts = decode_header(filename)
        decoded_string = ""
        for part, encoding in decoded_parts:
            if isinstance(part, bytes):
                decoded_string += part.decode(encoding or "utf-8", errors="replace")
            else:
                decoded_string += part
        return decoded_string
    except Exception:
        return filename

def is_allowed_file(filename: str, content_type: str) -> bool:
    allowed_mime_prefixes = ["image/"]
    allowed_mimes = {"application/pdf"}
    allowed_ext = {".pdf", ".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".tif", ".webp"}

    ext = os.path.splitext(filename.lower())[1]
    if ext in allowed_ext:
        return True
    if content_type in allowed_mimes:
        return True
    return any(content_type.startswith(prefix) for prefix in allowed_mime_prefixes)

def run_document_intelligence_ocr(file_bytes: bytes, content_type: str) -> str:
    url = (
        f"{Config.DOCINT_ENDPOINT}/documentintelligence/documentModels/"
        f"{Config.DOCINT_MODEL_ID}:analyze?api-version={Config.DOCINT_API_VERSION}"
    )
    headers = {
        "Ocp-Apim-Subscription-Key": Config.DOCINT_KEY,
        "Content-Type": content_type,
    }

    logging.info("Submitting file to Azure Document Intelligence OCR")
    response = requests.post(
        url,
        headers=headers,
        data=file_bytes,
        timeout=(Config.HTTP_CONNECT_TIMEOUT_SEC, Config.HTTP_READ_TIMEOUT_SEC),
    )

    if response.status_code not in (200, 202):
        raise APIError(
            "Document Intelligence analyze request failed.",
            502,
            {
                "status_code": response.status_code,
                "response": safe_json_or_text(response),
            },
        )

    operation_location = response.headers.get("Operation-Location")
    if not operation_location:
        payload = response.json() if response.content else {}
        status = str(payload.get("status", "")).lower()
        if status == "succeeded":
            text = (((payload.get("analyzeResult") or {}).get("content")) or "").strip()
            if not text:
                raise APIError("OCR completed but analyzeResult.content was empty.", 502)
            return text

        raise APIError(
            "Missing Operation-Location header from Document Intelligence response.",
            502,
            {"response": safe_json_or_text(response)},
        )

    return poll_docint_result(operation_location)


def poll_docint_result(operation_location: str) -> str:
    headers = {"Ocp-Apim-Subscription-Key": Config.DOCINT_KEY}

    interval = Config.DOCINT_POLL_INTERVAL_SEC
    for attempt in range(1, Config.DOCINT_POLL_MAX_RETRIES + 1):
        logging.info("Polling OCR operation (attempt %s/%s)", attempt, Config.DOCINT_POLL_MAX_RETRIES)

        response = requests.get(
            operation_location,
            headers=headers,
            timeout=(Config.HTTP_CONNECT_TIMEOUT_SEC, Config.HTTP_READ_TIMEOUT_SEC),
        )

        if response.status_code >= 400:
            raise APIError(
                "Failed while polling Document Intelligence operation.",
                502,
                {
                    "status_code": response.status_code,
                    "response": safe_json_or_text(response),
                    "operation_location": operation_location,
                },
            )

        payload = response.json()
        status = str(payload.get("status", "")).lower()

        if status == "succeeded":
            text = (((payload.get("analyzeResult") or {}).get("content")) or "").strip()
            if not text:
                raise APIError("OCR succeeded but analyzeResult.content was empty.", 502)
            return text

        if status in {"failed", "canceled", "cancelled"}:
            raise APIError(
                "Document Intelligence operation did not succeed.",
                502,
                {"status": status, "response": payload},
            )

        if attempt < Config.DOCINT_POLL_MAX_RETRIES:
            time.sleep(interval)
            interval = min(
                interval * Config.DOCINT_POLL_BACKOFF_MULTIPLIER,
                Config.DOCINT_POLL_MAX_INTERVAL_SEC,
            )

    raise APIError(
        "Document Intelligence polling timed out before completion.",
        504,
        {"max_retries": Config.DOCINT_POLL_MAX_RETRIES},
    )


def run_aoai_extraction(ocr_text: str, prompt: str = "") -> Dict[str, Any]:
    url = (
        f"{Config.AOAI_ENDPOINT}/openai/deployments/{Config.AOAI_DEPLOYMENT}/"
        f"chat/completions?api-version={Config.AOAI_API_VERSION}"
    )
    headers = {
        "api-key": Config.AOAI_KEY,
        "Content-Type": "application/json",
    }

    user_instruction = build_user_prompt(ocr_text=ocr_text, user_prompt=prompt)
    body = {
        "messages": [
            {"role": "system", "content": "Return JSON only"},
            {"role": "user", "content": user_instruction},
        ],
        "temperature": 0,
    }

    logging.info("Calling Azure OpenAI for structured extraction")
    response = requests.post(
        url,
        headers=headers,
        json=body,
        timeout=(Config.HTTP_CONNECT_TIMEOUT_SEC, Config.HTTP_READ_TIMEOUT_SEC),
    )

    if response.status_code >= 400:
        raise APIError(
            "Azure OpenAI request failed.",
            502,
            {
                "status_code": response.status_code,
                "response": safe_json_or_text(response),
            },
        )

    payload = response.json()
    content = extract_aoai_content(payload)
    parsed = parse_json_safely(content)

    if not isinstance(parsed, dict):
        raise APIError(
            "AI output must be a JSON object.",
            502,
            {"type": type(parsed).__name__},
        )

    return parsed

def build_user_prompt(ocr_text: str, user_prompt: str) -> str:
    base = """
            Extract structured data from the OCR text.

            Rules:
            - Keep values exactly as written in OCR
            - JSON keys should be simple Japanese words
            - Do not merge multiple labels into one field
            - If unclear, return ""
            - Return only valid JSON
            - DO NOT put important fields into 補足 if they can be inferred
            - Only use 補足 for truly irrelevant or unknown text
            - Keep all values as-is
            - Take all data OCR extracted
            """
    return base + "\nUser instructions:\n" + user_prompt + "\n\nOCR:\n" + ocr_text

def dict_to_csv_vertical(data: Dict[str, Any]) -> str:
    output = io.StringIO()

    writer = csv.writer(
        output,
        delimiter=",",
        quotechar='"',
        quoting=csv.QUOTE_MINIMAL,
        lineterminator="\r\n"
    )

    # ✅ write row-by-row (vertical)
    for key, value in data.items():
        writer.writerow([key, value])

    return output.getvalue()

def extract_aoai_content(payload: Dict[str, Any]) -> str:
    try:
        return payload["choices"][0]["message"]["content"]
    except (KeyError, IndexError, TypeError) as exc:
        raise APIError(
            "Unexpected Azure OpenAI response shape.",
            502,
            {"response": payload},
        ) from exc

def reorder_columns(data: Dict[str, Any]) -> list[str]:
    all_keys = list(data.keys())

    special_suffixes = ["原反", "加工賃"]

    special_keys = [
        k for k in all_keys
        if any(k.endswith(suffix) for suffix in special_suffixes)
    ]

    normal_keys = [k for k in all_keys if k not in special_keys]

    return normal_keys + special_keys

def flatten_json(data: Dict[str, Any], parent_key: str = "", sep: str = "_") -> Dict[str, Any]:
    items: Dict[str, Any] = {}

    for k, v in data.items():
        k = clean_key(k)
        new_key = f"{parent_key}{sep}{k}" if parent_key else k

        if isinstance(v, dict):
            items.update(flatten_json(v, new_key, sep=sep))
        elif isinstance(v, list):
            # nếu list là danh sách object, gộp JSON string vào 1 cell
            items[new_key] = json.dumps(v, ensure_ascii=False)
        else:
            items[new_key] = v

    return items

def clean_key(key: str) -> str:
    return (
        key.replace('"', '')
           .replace("'", "")
           .replace(" ", "")
           .replace("\n", "")
           .replace("\t", "")
           .strip()
    )

def split_summary_and_tables(data: dict, parent_key: str = ""):
    summary = {}
    tables = {}

    for k, v in data.items():
        key = f"{parent_key}_{k}" if parent_key else k

        if isinstance(v, dict):
            child_summary, child_tables = split_summary_and_tables(v, key)
            summary.update(child_summary)
            tables.update(child_tables)

        elif isinstance(v, list) and v and all(isinstance(i, dict) for i in v):
            tables[key] = v

        else:
            summary[key] = v

    return summary, tables

def write_vertical_summary(ws, summary: dict, start_row=1):
    row = start_row
    for k, v in summary.items():
        ws.cell(row=row, column=1, value=k)
        ws.cell(row=row, column=2, value=v)
        row += 1
    return row

def write_table(ws, rows: list[dict], start_row: int):
    if not rows:
        return start_row

    headers = list(rows[0].keys())

    # ưu tiên 原反 / 加工賃 về bên phải
    special = ["原反", "加工賃"]
    normal_headers = [h for h in headers if h not in special]
    special_headers = [h for h in special if h in headers]
    headers = normal_headers + special_headers

    # header row
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F6D8C")
        cell.alignment = Alignment(horizontal="center")

    # data rows
    for r_idx, item in enumerate(rows, start=start_row + 1):
        for c_idx, h in enumerate(headers, start=1):
            value = item.get(h, "")

            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)

            ws.cell(row=r_idx, column=c_idx, value=value)
            
    return start_row + len(rows) + 2

def build_xlsx_from_ai_result(ai_result: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "結果"

    summary, tables = split_summary_and_tables(ai_result)

    current_row = 1
    current_row = write_vertical_summary(ws, summary, start_row=current_row)
    current_row += 2

    for table_name, rows in tables.items():
        ws.cell(row=current_row, column=1, value=table_name)
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1
        current_row = write_table(ws, rows, start_row=current_row)

    # auto width đơn giản
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def parse_json_safely(content: str) -> Any:
    try:
        return json.loads(content)
    except json.JSONDecodeError:
        pass

    cleaned = strip_markdown_fences(content)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass

    decoder = json.JSONDecoder()
    for i, ch in enumerate(cleaned):
        if ch not in "[{":
            continue
        try:
            obj, _ = decoder.raw_decode(cleaned[i:])
            return obj
        except json.JSONDecodeError:
            continue

    raise APIError("AI response was not valid JSON.", 502, {"raw": content[:2000]})


def strip_markdown_fences(text: str) -> str:
    stripped = text.strip()
    if stripped.startswith("```") and stripped.endswith("```"):
        lines = stripped.splitlines()
        if len(lines) >= 3:
            return "\n".join(lines[1:-1]).strip()
    return stripped


def safe_json_or_text(response: RequestsResponse) -> Any:
    try:
        return response.json()
    except ValueError:
        return response.text[:4000]


if __name__ == "__main__":
    port = int(os.getenv("PORT", "8000"))
    uvicorn.run("app:app", host="127.0.0.1", port=port, reload=True, http="httptools")
